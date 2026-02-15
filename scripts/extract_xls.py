#!/usr/bin/env python3
"""Extract text from an XLSX file (first/last N rows).

Outputs normalized text (lowercased, whitespace-collapsed) to stdout.

Exit codes:
    0 - success
    1 - extraction produced empty text
    2 - error (file not found, unreadable, etc.)
"""

import argparse
import re
import sys

from openpyxl import load_workbook


def normalize(text: str) -> str:
    """Lowercase and collapse all whitespace to single spaces."""
    return re.sub(r"\s+", " ", text.lower()).strip()


def row_to_text(row) -> str:
    """Convert a row of cells to a space-separated string."""
    parts = []
    for cell in row:
        if cell.value is not None:
            parts.append(str(cell.value))
    return " ".join(parts)


def extract(xlsx_path: str, lines: int) -> str:
    """Return normalized text from first N rows + last N rows of the first sheet."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    # Collect all rows as text lines
    all_rows = []
    for row in ws.iter_rows():
        text = row_to_text(row)
        if text.strip():
            all_rows.append(text)

    wb.close()

    if not all_rows:
        return ""

    if len(all_rows) <= lines * 2:
        combined = all_rows
    else:
        combined = all_rows[:lines] + all_rows[-lines:]

    raw = "\n".join(combined)
    return normalize(raw)


def main():
    parser = argparse.ArgumentParser(description="Extract text from an XLSX file.")
    parser.add_argument("file", help="Path to XLSX file")
    parser.add_argument(
        "--lines",
        type=int,
        default=15,
        help="Number of rows to extract from first/last section (default: 15)",
    )
    args = parser.parse_args()

    try:
        text = extract(args.file, args.lines)
    except Exception as e:
        print(f"Error extracting XLSX: {e}", file=sys.stderr)
        sys.exit(2)

    if not text.strip():
        print("Error: extraction produced empty text", file=sys.stderr)
        sys.exit(1)

    print(text)


if __name__ == "__main__":
    main()
